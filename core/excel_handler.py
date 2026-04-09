"""
Gestione Excel: creazione e lettura del file movimenti.
"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COL = dict(data_op=0, dare=1, avere=2, descrizione=3, causale=4,
           conto_dare=5, conto_avere=6, numero_doc=7, data_reg=8, da_importare=9)

def _fmt(v):
    if v is None: return "-"
    return f"€ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

def _is_date(val):
    if val is None: return False
    if isinstance(val, datetime): return True
    for fmt in ['%Y-%m-%d','%d/%m/%Y','%d/%m/%y']:
        try: datetime.strptime(str(val).strip(), fmt); return True
        except: pass
    return False

def _dstr(val):
    if val is None: return ''
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    for fmt in ['%Y-%m-%d','%d/%m/%Y','%d/%m/%y']:
        try: return datetime.strptime(str(val).strip(), fmt).strftime('%Y-%m-%d')
        except: pass
    return str(val)

def _flt(val):
    if val is None: return None
    try: return float(val)
    except: return None

def _str(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s.lower() in ('none','nan') else s

def create_xlsx_bytes(transactions, saldo_iniziale, banca_nome=""):
    """
    Crea il file Excel e lo restituisce come bytes (per download Streamlit).
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Movimenti"

    hf  = PatternFill("solid", fgColor="1A3A6E")
    hft = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    tf  = PatternFill("solid", fgColor="2E5090")
    brd = Border(
        left=Side(style='thin', color='C0C8D8'),
        right=Side(style='thin', color='C0C8D8'),
        top=Side(style='thin', color='C0C8D8'),
        bottom=Side(style='thin', color='C0C8D8')
    )

    label = f"ESTRATTO CONTO → PRIMA NOTA  |  {banca_nome}" if banca_nome else "ESTRATTO CONTO → PRIMA NOTA"
    ws['A1'] = label
    ws['A1'].font = Font(bold=True, size=13, color="1A3A6E", name="Calibri")
    ws.merge_cells('A1:J1')

    si_str = _fmt(saldo_iniziale) if saldo_iniziale is not None else "N/D"
    ws['A2'] = f"Saldo iniziale periodo: {si_str}"
    ws['A2'].font = Font(name="Calibri", size=9, italic=True, color="444466")
    ws.merge_cells('A2:J2')

    ws['A3'] = "🟢 DARE az. = entrate nel c/c   🔴 AVERE az. = uscite dal c/c"
    ws['A3'].font = Font(name="Calibri", size=8, italic=True, color="555500")
    ws.merge_cells('A3:J3')

    headers    = ["Data Op.","DARE az. (€)","AVERE az. (€)","Descrizione",
                  "Causale *","Conto Dare *","Conto Avere *","N° Documento",
                  "Data Reg. *","Da importare"]
    col_widths = [12, 14, 14, 60, 14, 14, 14, 18, 12, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = hf; c.font = hft; c.border = brd
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "A5"

    ef  = PatternFill("solid", fgColor="EAF5EA")
    ef2 = PatternFill("solid", fgColor="D8EED8")
    uf  = PatternFill("solid", fgColor="FAF0F0")
    uf2 = PatternFill("solid", fgColor="F5E0E0")
    amt = '#,##0.00'

    for i, tx in enumerate(transactions, 5):
        dare_az  = tx.get('entrata')
        avere_az = tx.get('uscita')
        is_e = dare_az is not None
        fill = (ef if i % 2 == 0 else ef2) if is_e else (uf if i % 2 == 0 else uf2)
        vals = [
            tx['data_op'], dare_az, avere_az, tx['descrizione'],
            tx.get('causale', ''), tx.get('conto_dare', ''), tx.get('conto_avere', ''),
            tx.get('numero_doc', ''), tx['data_op'], 'SI'
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=9); c.border = brd; c.fill = fill
            if col in [2, 3] and val is not None:
                c.number_format = amt; c.alignment = Alignment(horizontal='right')
            elif col == 4:
                c.alignment = Alignment(wrap_text=True, vertical='top')

    # Riga TOTALI
    tr = len(transactions) + 5
    ws.cell(row=tr, column=4, value="◀  TOTALI  ▶").font = Font(bold=True, color="FFFFFF", name="Calibri")
    for col in [2, 3]:
        ltr = openpyxl.utils.get_column_letter(col)
        c = ws.cell(row=tr, column=col, value=f"=SUM({ltr}5:{ltr}{tr-1})")
        c.number_format = amt
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for col in range(1, 11):
        ws.cell(row=tr, column=col).fill = tf
        ws.cell(row=tr, column=col).border = brd
        if col not in [2, 3, 4]:
            ws.cell(row=tr, column=col).font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def read_xlsx_bytes(file_bytes_or_path):
    """
    Legge il file Excel (bytes o path) e restituisce lista transazioni.
    """
    if isinstance(file_bytes_or_path, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes_or_path), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)

    ws = wb.active
    txs = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not _is_date(row[COL['data_op']]): continue
        if _str(row[COL['da_importare']]).upper() == 'NO': continue
        txs.append({
            'data_op'    : _dstr(row[COL['data_op']]),
            'dare_az'    : _flt(row[COL['dare']]),
            'avere_az'   : _flt(row[COL['avere']]),
            'descrizione': _str(row[COL['descrizione']]),
            'causale'    : _str(row[COL['causale']]),
            'conto_dare' : _str(row[COL['conto_dare']]),
            'conto_avere': _str(row[COL['conto_avere']]),
            'numero_doc' : _str(row[COL['numero_doc']]),
            'data_reg'   : _dstr(row[COL['data_reg']]) or _dstr(row[COL['data_op']]),
        })
    return txs
